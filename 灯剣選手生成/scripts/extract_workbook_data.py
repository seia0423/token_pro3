import json
import os
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


SOURCE = r"C:\Users\Owner\OneDrive\デスクトップ\灯剣選手生成完成版.xlsm"
OUTPUT = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "generator-data.json")
JS_OUTPUT = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "generator-data.js")


def compact(values):
    return [value for value in values if value not in (None, "")]


def read_column(ws, col_letter, start_row=2):
    col = column_index_from_string(col_letter)
    items = []
    seen = set()
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row, col).value
        if value in (None, ""):
            continue
        text = str(value).strip()
        if text and text not in seen:
            seen.add(text)
            items.append(text)
    return items


def main():
    wb = load_workbook(SOURCE, data_only=True)
    formula_wb = load_workbook(SOURCE, data_only=False)

    generator = wb["生成機"]
    names_ws = wb["名前リスト"]
    nationalities_ws = wb["国籍リスト"]
    mapping_ws = wb["国籍名前対照"]
    skills_ws = wb["特能"]

    stat_names = [generator.cell(1, col).value for col in range(11, 41)]
    position_codes = [str(generator.cell(1, col).value) for col in range(41, 57)]
    position_weights = {}
    for row in range(41, 57):
        code = str(generator.cell(row, 10).value)
        weights = [generator.cell(row, col).value or 0 for col in range(11, 41)]
        position_weights[code] = weights

    nationality_groups = {}
    for col in range(1, nationalities_ws.max_column + 1):
        header = nationalities_ws.cell(1, col).value
        if header in (None, ""):
            continue
        values = compact(nationalities_ws.cell(row, col).value for row in range(2, nationalities_ws.max_row + 1))
        if values:
            nationality_groups[str(header)] = [str(value) for value in values]

    name_column_map = {}
    for row in range(2, mapping_ws.max_row + 1):
        nationality = mapping_ws.cell(row, 1).value
        gender = mapping_ws.cell(row, 2).value
        surname_col = mapping_ws.cell(row, 3).value
        given_col = mapping_ws.cell(row, 4).value
        if all(value not in (None, "") for value in [nationality, gender, surname_col, given_col]):
            name_column_map[f"{nationality}|{gender}"] = {
                "surnameColumn": str(surname_col),
                "givenNameColumn": str(given_col),
            }

    names_by_key = {}
    for key, columns in name_column_map.items():
        names_by_key[key] = {
            "surnames": read_column(names_ws, columns["surnameColumn"]),
            "givenNames": read_column(names_ws, columns["givenNameColumn"]),
        }

    skill_headers = [skills_ws.cell(1, col).value for col in range(1, skills_ws.max_column + 1)]
    skills = []
    for row in range(2, skills_ws.max_row + 1):
        name = skills_ws.cell(row, 1).value
        if name in (None, ""):
            continue
        entry = {"name": str(name), "modifiers": {}, "note": skills_ws.cell(row, 28).value or ""}
        for col in range(2, 28):
            value = skills_ws.cell(row, col).value
            if isinstance(value, (int, float)) and value:
                entry["modifiers"][str(skill_headers[col - 1])] = value
        for col in range(29, 34):
            value = skills_ws.cell(row, col).value
            if isinstance(value, (int, float)) and value:
                entry.setdefault("positionBands", {})[str(skill_headers[col - 1])] = value
        skills.append(entry)

    formulas = {}
    formula_sheet = formula_wb["生成機"]
    for addr in ["F2", "G2", "H2", "I2", "J2", "K2", "AO2"]:
        formulas[addr] = formula_sheet[addr].value

    data = {
        "sourceWorkbook": SOURCE,
        "statNames": stat_names,
        "positionCodes": position_codes,
        "positionWeights": position_weights,
        "nationalityGroups": nationality_groups,
        "nameColumnMap": name_column_map,
        "namesByKey": names_by_key,
        "skills": skills,
        "referenceFormulas": formulas,
    }

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    with open(JS_OUTPUT, "w", encoding="utf-8") as f:
        f.write("window.TOUKEN_GENERATOR_DATA = ")
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write(";\n")
    print(OUTPUT)


if __name__ == "__main__":
    main()
