import json
import os
import sys
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


PATH = r"C:\Users\Owner\OneDrive\デスクトップ\灯剣選手生成完成版.xlsm"


def non_empty_cells(ws, limit=40):
    rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                rows.append(
                    {
                        "cell": cell.coordinate,
                        "value": str(cell.value)[:160],
                        "data_type": cell.data_type,
                    }
                )
                if len(rows) >= limit:
                    return rows
    return rows


def formula_cells(ws, limit=60):
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append({"cell": cell.coordinate, "formula": cell.value[:240]})
                if len(formulas) >= limit:
                    return formulas
    return formulas


def workbook_xml_summary():
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    with zipfile.ZipFile(PATH) as zf:
        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = []
        for sheet in wb.findall(".//main:sheet", ns):
            sheets.append(
                {
                    "name": sheet.attrib.get("name"),
                    "sheetId": sheet.attrib.get("sheetId"),
                    "state": sheet.attrib.get("state", "visible"),
                    "rid": sheet.attrib.get("{%s}id" % ns["rel"]),
                }
            )
        names = []
        for node in wb.findall(".//main:definedName", ns):
            names.append({"name": node.attrib.get("name"), "text": (node.text or "")[:240]})
        entries = zf.namelist()
    return {"sheets_xml": sheets, "defined_names": names, "zip_entries": entries}


def main():
    wb = load_workbook(PATH, keep_vba=True, data_only=False)
    data_wb = load_workbook(PATH, keep_vba=True, data_only=True)
    if "--summary" in sys.argv:
        for ws in wb.worksheets:
            formula_count = 0
            value_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        value_count += 1
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
            headers = ",".join(str(c.value) for c in ws[1] if c.value is not None)
            print(
                f"{ws.title}\t{ws.max_row}x{ws.max_column}\t"
                f"values={value_count}\tformulas={formula_count}\theaders={headers[:240]}"
            )
        return
    if "--generator-ranges" in sys.argv:
        ws = wb["生成機"]
        print("B20:B24", [ws[f"B{i}"].value for i in range(20, 25)])
        print("headers", [ws.cell(1, c).value for c in range(1, ws.max_column + 1)])
        for r in range(40, 57):
            print(r, [ws.cell(r, c).value for c in range(1, 57)])
        return
    if "--key-formulas" in sys.argv:
        for sheet_name in ["生成機", "㊙"]:
            ws = wb[sheet_name]
            print(f"[{sheet_name}]")
            for addr in ["F2", "G2", "K2", "AO2", "BE2", "BF2", "BG2", "BH2"]:
                print(addr)
                print(ws[addr].value)
        return
    report = {
        "file": PATH,
        "size_bytes": os.path.getsize(PATH),
        "workbook": workbook_xml_summary(),
        "sheets": [],
    }

    for ws in wb.worksheets:
        data_ws = data_wb[ws.title]
        formula_count = 0
        formula_examples = []
        headers = []
        for cell in ws[1]:
            if cell.value is not None:
                headers.append({"cell": cell.coordinate, "value": str(cell.value)})
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    if len(formula_examples) < 12:
                        formula_examples.append({"cell": cell.coordinate, "formula": cell.value[:180]})
        report["sheets"].append(
            {
                "title": ws.title,
                "state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "headers": headers[:80],
                "formula_count": formula_count,
                "formula_examples": formula_examples,
                "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:20]],
                "tables": list(ws.tables.keys()),
                "non_empty_sample": non_empty_cells(ws, limit=15),
                "data_sample": non_empty_cells(data_ws, limit=10),
            }
        )

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
